from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .models import ExtractionResult, TagRecord


HEADERS = [
    "Tag",
    "Normalized Tag",
    "Tag Type",
    "P&ID Number",
    "Page",
    "Source",
    "Confidence",
    "BBox",
    "Context",
    "MDS Status",
    "MDS Asset ID",
    "MDS Description",
    "MDS Discipline",
    "MDS System",
    "MDS Criticality",
    "MDS Message",
]


def export_results_to_excel(result: ExtractionResult, path: Path) -> None:
    workbook = Workbook()
    tags_sheet = workbook.active
    tags_sheet.title = "Tags"
    _write_tags(tags_sheet, result.tags)
    _write_summary(workbook.create_sheet("Summary"), result)
    _write_validation(workbook.create_sheet("MDS Validation"), result.tags)
    _write_raw_text(workbook.create_sheet("Raw Text"), result)
    workbook.save(path)


def _write_tags(sheet: Worksheet, tags: list[TagRecord]) -> None:
    sheet.append(HEADERS)
    for tag in tags:
        sheet.append(
            [
                tag.tag,
                tag.normalized_tag,
                tag.tag_type,
                tag.pid_number,
                tag.page,
                tag.source,
                tag.confidence,
                tag.bbox,
                tag.context,
                tag.mds_status,
                tag.mds_asset_id,
                tag.mds_description,
                tag.mds_discipline,
                tag.mds_system,
                tag.mds_criticality,
                tag.mds_message,
            ]
        )
    _format_table(sheet)


def _write_summary(sheet: Worksheet, result: ExtractionResult) -> None:
    sheet.append(["Metric", "Value"])
    sheet.append(["Source PDF", result.source_pdf.name])
    sheet.append(["P&ID Number", result.pid_number or "Not detected"])
    sheet.append(["Total Tag Rows", len(result.tags)])
    sheet.append(["Unique Tags", len({tag.normalized_tag for tag in result.tags})])
    sheet.append(["Warnings", " | ".join(result.warnings) if result.warnings else "None"])
    sheet.append([])
    sheet.append(["Tag Type", "Count"])
    for key, count in Counter(tag.tag_type for tag in result.tags).most_common():
        sheet.append([key, count])
    sheet.append([])
    sheet.append(["MDS Status", "Count"])
    for key, count in Counter(tag.mds_status for tag in result.tags).most_common():
        sheet.append([key, count])
    _format_table(sheet)


def _write_validation(sheet: Worksheet, tags: list[TagRecord]) -> None:
    sheet.append(["Normalized Tag", "MDS Status", "Asset ID", "Description", "System", "Criticality", "Message"])
    seen: set[str] = set()
    for tag in tags:
        if tag.normalized_tag in seen:
            continue
        seen.add(tag.normalized_tag)
        sheet.append(
            [
                tag.normalized_tag,
                tag.mds_status,
                tag.mds_asset_id,
                tag.mds_description,
                tag.mds_system,
                tag.mds_criticality,
                tag.mds_message,
            ]
        )
    _format_table(sheet)


def _write_raw_text(sheet: Worksheet, result: ExtractionResult) -> None:
    sheet.append(["Page", "Source", "Text"])
    for page in result.raw_pages:
        sheet.append([page.page, page.source, page.text[:32000]])
    _format_table(sheet)
    sheet.column_dimensions["C"].width = 120


def _format_table(sheet: Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        letter = column_cells[0].column_letter
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

